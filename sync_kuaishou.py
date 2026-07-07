import argparse
import csv
import json
import msvcrt
import os
import sqlite3
import subprocess
import sys
import time
from contextlib import redirect_stderr, redirect_stdout
from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from kuaishou_client import KuaishouApiError, KuaishouClient, load_runtime_env
from kuaishou_store import KuaishouStore


ROOT_DIR = Path(__file__).resolve().parent
DEMAND_EXCEL = ROOT_DIR / "需求账号.xlsx"
LOG_DIR = ROOT_DIR / "logs"
SCHEDULE_TASK_NAME = "快手数据库监控_30分钟拉数"
POWERSHELL_EXE = Path(os.environ.get("SystemRoot", r"C:\Windows")) / "System32" / "WindowsPowerShell" / "v1.0" / "powershell.exe"
TOKEN_ALERT_COOLDOWN_SECONDS = 6 * 60 * 60
SCHEDULE_LOCK_FILE = LOG_DIR / "scheduled_pull.lock"
DEFAULT_DEMAND_ACCOUNT = {
    "启用": "是",
    "广告主ID": 39059876,
    "广告主名称": "大师爱闯关",
    "备注": "初始有消耗账号",
}
SHANGHAI_TZ = timezone(timedelta(hours=8), name="Asia/Shanghai")

CAMPAIGN_LIST = "/rest/openapi/gw/dsp/campaign/list"
UNIT_LIST = "/rest/openapi/gw/dsp/unit/list"
CREATIVE_LIST = "/rest/openapi/gw/dsp/creative/list"
ADVERTISER_LIST = "/rest/openapi/gw/uc/v1/advertisers"
ACCOUNT_REPORT = "/rest/openapi/v1/report/account_report"
UNIT_REPORT = "/rest/openapi/v1/report/unit_report"


def print_json(value):
    sys.stdout.buffer.write((json.dumps(value, ensure_ascii=False, indent=2) + "\n").encode("utf-8"))


def log_line(message):
    stamp = shanghai_now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{stamp}] {message}", flush=True)


def shanghai_now():
    return datetime.now(SHANGHAI_TZ)


def is_token_error(error):
    text = str(error)
    body = getattr(error, "body", None)
    if isinstance(body, dict):
        text = f"{text} {body.get('message', '')} {body.get('code', '')}"
    lowered = text.lower()
    return "token" in lowered and ("错误" in text or "invalid" in lowered or "expired" in lowered or "402005" in text)


def maybe_alert_token_problem(message, force=False):
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    now = shanghai_now()
    state_path = LOG_DIR / "token_alert_state.json"
    alert_path = LOG_DIR / f"token_alert_{now.strftime('%Y%m%d')}.txt"
    state = {}
    if state_path.exists():
        try:
            state = json.loads(state_path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            state = {}

    last_alert = 0.0
    try:
        last_alert = float(state.get("last_alert_ts") or 0)
    except (TypeError, ValueError):
        last_alert = 0.0
    should_notify = force or time.time() - last_alert >= TOKEN_ALERT_COOLDOWN_SECONDS

    line = f"[{now.strftime('%Y-%m-%d %H:%M:%S')}] {message}\n"
    alert_path.write_text((alert_path.read_text(encoding="utf-8") if alert_path.exists() else "") + line, encoding="utf-8")

    if should_notify:
        state_path.write_text(
            json.dumps(
                {
                    "last_alert_ts": time.time(),
                    "last_alert_at": now.strftime("%Y-%m-%d %H:%M:%S"),
                    "message": message,
                    "alert_file": str(alert_path),
                },
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )
        notify_text = f"快手数据监控 token 异常：{message}。请查看 {alert_path}"
        try:
            subprocess.run(["msg.exe", "*", "/TIME:60", notify_text], check=False, capture_output=True, text=True)
        except Exception:
            pass
    return {"alert_file": str(alert_path), "notified": should_notify}


def redact_sensitive(value):
    if isinstance(value, list):
        return [redact_sensitive(item) for item in value]
    if not isinstance(value, dict):
        return value

    redacted = {}
    for key, item in value.items():
        lower_key = key.lower()
        if "token" in lower_key or lower_key in {"secret", "app_id"}:
            redacted[key] = "<hidden>" if item else item
        else:
            redacted[key] = redact_sensitive(item)
    return redacted


def create_store(args, init_schema=True, set_wal=True):
    store = KuaishouStore(args.db, set_wal=set_wal)
    try:
        if init_schema:
            store.init_schema()
        return store
    except Exception:
        store.close()
        raise


def is_sqlite_locked(error):
    return isinstance(error, sqlite3.OperationalError) and "locked" in str(error).lower()


def create_store_with_retry(args, init_schema=True, set_wal=True, attempts=6, delay_seconds=10):
    last_error = None
    for attempt in range(1, attempts + 1):
        try:
            return create_store(args, init_schema=init_schema, set_wal=set_wal)
        except sqlite3.OperationalError as error:
            last_error = error
            if not is_sqlite_locked(error) or attempt >= attempts:
                raise
            log_line(
                f"database locked while opening store attempt={attempt}/{attempts}; "
                f"retry in {delay_seconds}s"
            )
            time.sleep(delay_seconds)
    raise last_error


def run_sqlite_write_with_retry(description, operation, attempts=6, delay_seconds=10):
    last_error = None
    for attempt in range(1, attempts + 1):
        try:
            return operation()
        except sqlite3.OperationalError as error:
            last_error = error
            if not is_sqlite_locked(error) or attempt >= attempts:
                raise
            log_line(
                f"database locked during {description} attempt={attempt}/{attempts}; "
                f"retry in {delay_seconds}s"
            )
            time.sleep(delay_seconds)
    raise last_error


def try_acquire_file_lock(path):
    path.parent.mkdir(parents=True, exist_ok=True)
    handle = path.open("a+")
    try:
        handle.seek(0)
        msvcrt.locking(handle.fileno(), msvcrt.LK_NBLCK, 1)
        handle.seek(0)
        handle.truncate()
        handle.write(f"pid={os.getpid()} started_at={shanghai_now().isoformat()}\n")
        handle.flush()
        return handle
    except OSError:
        handle.close()
        return None


def release_file_lock(handle):
    if not handle:
        return
    try:
        handle.seek(0)
        msvcrt.locking(handle.fileno(), msvcrt.LK_UNLCK, 1)
    finally:
        handle.close()


def command_init_db(args):
    store = create_store_with_retry(args, init_schema=True, attempts=6, delay_seconds=10)
    try:
        print_json({"ok": True, "db": str(store.db_path)})
    finally:
        store.close()


def command_status(args):
    client = KuaishouClient()
    print_json(client.token_status())


def command_exchange(args):
    client = KuaishouClient()
    print_json(redact_sensitive(client.exchange_access_token(args.auth_code)))


def command_refresh(args):
    client = KuaishouClient()
    print_json(redact_sensitive(client.refresh_access_token()))


def command_check_token(args):
    client = KuaishouClient()
    try:
        result = client.refresh_access_token()
        print_json({"ok": True, "message": "token refresh ok", "result": redact_sensitive(result)})
    except Exception as error:
        alert = maybe_alert_token_problem(f"手动检测 token 失败：{error}", force=True)
        if isinstance(error, KuaishouApiError):
            print_json({"ok": False, "error": str(error), "status": error.status, "detail": error.body, "alert": alert})
        else:
            print_json({"ok": False, "error": str(error), "alert": alert})
        raise SystemExit(1)


def command_init_demand_excel(args):
    path = ensure_demand_excel()
    print_json({"ok": True, "file": str(path), "sheet": "需求账号"})


def command_automation_status(args):
    from automation_center import AutomationCenter

    center = AutomationCenter(args.db)
    try:
        dashboard = center.dashboard()
        print_json(dashboard)
    finally:
        center.close()


def command_automation_run(args):
    from automation_center import AutomationCenter

    center = AutomationCenter(args.db)
    try:
        result = center.run_rules(
            source="cli",
            dry_run=not args.execute,
            respect_enabled=not args.ignore_disabled,
        )
        print_json(result)
    finally:
        center.close()


def command_web(args):
    from automation_web import run_server

    run_server(host=args.host, port=args.port, db_path=args.db)


def command_scheduled_pull(args):
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    log_file = LOG_DIR / f"scheduled_pull_{shanghai_now().strftime('%Y%m%d')}.log"
    failed = False
    with log_file.open("a", encoding="utf-8") as handle:
        with redirect_stdout(handle), redirect_stderr(handle):
            lock_handle = try_acquire_file_lock(SCHEDULE_LOCK_FILE)
            if not lock_handle:
                result = {
                    "ok": True,
                    "skipped": True,
                    "reason": "previous scheduled-pull is still running",
                    "finished_at": shanghai_now().isoformat(),
                }
                log_line(f"scheduled-pull skipped reason={result['reason']}")
                print_json(result)
                return
            try:
                try:
                    result = run_scheduled_pull(args, log_file)
                except Exception as error:
                    failed = True
                    result = {
                        "ok": False,
                        "summary": {
                            "log_file": str(log_file),
                            "error_type": type(error).__name__,
                            "error": str(error),
                            "finished_at": shanghai_now().isoformat(),
                        },
                    }
                    log_line(f"scheduled-pull failed error_type={type(error).__name__} error={error}")
                print_json(result)
            finally:
                release_file_lock(lock_handle)
    if failed:
        raise SystemExit(1)


def command_install_schedule(args):
    task_argument = f'"{ROOT_DIR / "sync_kuaishou.py"}" scheduled-pull'
    ps_script = f"""
$ErrorActionPreference = 'Stop'
$Action = New-ScheduledTaskAction -Execute {powershell_quote(sys.executable)} -Argument {powershell_quote(task_argument)} -WorkingDirectory {powershell_quote(str(ROOT_DIR))}
$Trigger = New-ScheduledTaskTrigger -Once -At (Get-Date).Date -RepetitionInterval (New-TimeSpan -Minutes 30) -RepetitionDuration (New-TimeSpan -Days 3650)
$Settings = New-ScheduledTaskSettingsSet -AllowStartIfOnBatteries -DontStopIfGoingOnBatteries -MultipleInstances IgnoreNew
Register-ScheduledTask -TaskName {powershell_quote(SCHEDULE_TASK_NAME)} -Action $Action -Trigger $Trigger -Settings $Settings -Description '快手数据库监控每30分钟拉取需求账号数据' -Force | Out-Null
"""
    subprocess.run(
        [str(POWERSHELL_EXE), "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", ps_script],
        check=True,
        text=True,
        errors="replace",
        capture_output=True,
    )
    query = subprocess.run(
        ["schtasks.exe", "/Query", "/TN", SCHEDULE_TASK_NAME],
        check=True,
        text=True,
        errors="replace",
        capture_output=True,
    )
    print_json(
        {
            "ok": True,
            "task_name": SCHEDULE_TASK_NAME,
            "execute": sys.executable,
            "arguments": task_argument,
            "working_directory": str(ROOT_DIR),
            "query": query.stdout.strip(),
        }
    )


def command_accounts(args):
    client = KuaishouClient()
    store = create_store(args)
    run_id = store.start_run("accounts")
    try:
        auth_user_id = args.auth_user_id or os.getenv("KUAISHOU_AUTH_USER_ID")
        if not auth_user_id:
            raise SystemExit("请提供 --auth-user-id，或在 .env 中配置 KUAISHOU_AUTH_USER_ID")
        result = client.request(
            ADVERTISER_LIST,
            method="POST",
            body={"advertiser_id": int(auth_user_id)},
        )
        client._assert_success(result)
        store.save_raw_response(run_id, ADVERTISER_LIST, {"advertiser_id": auth_user_id}, result)
        details = ((result.get("data") or {}).get("details") or [])
        count = store.upsert_advertisers(details, run_id)
        stats = {"advertisers": count}
        store.finish_run(run_id, "success", stats)
        print_json({"ok": True, "run_id": run_id, "stats": stats})
    except Exception as error:
        store.finish_run(run_id, "failed", message=str(error))
        raise
    finally:
        store.close()


def command_sync(args):
    advertiser_ids = resolve_advertiser_ids(args)
    if not advertiser_ids:
        raise SystemExit("请提供 --advertiser-id，或在 .env 中配置 KUAISHOU_ADVERTISER_ID")

    client = KuaishouClient()
    store = create_store(args)
    all_stats = {}
    try:
        for advertiser_id in advertiser_ids:
            run_id = store.start_run("sync", advertiser_id)
            try:
                stats = sync_advertiser(client, store, run_id, advertiser_id, args)
                store.finish_run(run_id, "success", stats)
                all_stats[str(advertiser_id)] = {"run_id": run_id, **stats}
            except Exception as error:
                store.finish_run(run_id, "failed", message=str(error))
                all_stats[str(advertiser_id)] = {"run_id": run_id, "error": str(error)}
                if not args.continue_on_error:
                    raise
        print_json({"ok": True, "db": str(store.db_path), "advertisers": all_stats})
    finally:
        store.close()


def command_snapshot(args):
    client = KuaishouClient()
    store = create_store(args)
    advertiser_id = int(args.advertiser_id)
    campaign_id = int(args.campaign_id)
    run_id = store.start_run("snapshot", advertiser_id)
    try:
        stats = sync_advertiser(client, store, run_id, advertiser_id, args, campaign_id=campaign_id)
        store.finish_run(run_id, "success", stats)
        print_json({"ok": True, "run_id": run_id, "stats": stats})
    except Exception as error:
        store.finish_run(run_id, "failed", message=str(error))
        raise
    finally:
        store.close()


def command_report(args):
    advertiser_ids = resolve_advertiser_ids(args)
    if not advertiser_ids:
        raise SystemExit("请提供 --advertiser-id，或在 .env 中配置 KUAISHOU_ADVERTISER_ID")

    client = KuaishouClient()
    store = create_store(args)
    all_stats = {}
    try:
        for advertiser_id in advertiser_ids:
            run_id = store.start_run("report", advertiser_id)
            try:
                stats = sync_report_range(client, store, run_id, advertiser_id, args)
                store.finish_run(run_id, "success", stats)
                all_stats[str(advertiser_id)] = {"run_id": run_id, **stats}
            except Exception as error:
                store.finish_run(run_id, "failed", message=str(error))
                all_stats[str(advertiser_id)] = {"run_id": run_id, "error": str(error)}
                if not args.continue_on_error:
                    raise
        print_json({"ok": True, "db": str(store.db_path), "advertisers": all_stats})
    finally:
        store.close()


def command_import_report_csv(args):
    store = create_store(args)
    advertiser_id = int(args.advertiser_id)
    source_file = str(Path(args.file).resolve())
    run_id = store.start_run("report_csv", advertiser_id)
    try:
        with open(args.file, "r", encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.DictReader(handle))
        count = store.upsert_daily_reports(
            advertiser_id,
            rows,
            run_id,
            source="CSV导入",
            source_file=source_file,
        )
        stats = {"快手日报": count, "来源文件": source_file}
        store.finish_run(run_id, "success", stats)
        print_json({"ok": True, "run_id": run_id, "stats": stats})
    except Exception as error:
        store.finish_run(run_id, "failed", message=str(error))
        raise
    finally:
        store.close()


def command_spenders(args):
    store = create_store(args)
    client = KuaishouClient()
    rows = store.conn.execute(
        """
        SELECT advertiser_id
        FROM advertisers
        ORDER BY advertiser_id
        """
    ).fetchall()
    advertiser_ids = [int(row["advertiser_id"]) for row in rows]
    if args.limit:
        advertiser_ids = advertiser_ids[: args.limit]
    if not advertiser_ids:
        raise SystemExit("当前 advertisers 表为空，请先运行 python .\\sync_kuaishou.py accounts")

    summary = {
        "checked_accounts": 0,
        "spend_accounts": 0,
        "synced_accounts": 0,
        "errors": 0,
        "spenders": [],
    }
    try:
        for index, advertiser_id in enumerate(advertiser_ids, start=1):
            if index > 1 and args.request_delay > 0:
                time.sleep(args.request_delay)
            print(f"[{index}/{len(advertiser_ids)}] 检查广告主 {advertiser_id}", file=sys.stderr)
            report_run_id = store.start_run("spenders_report", advertiser_id)
            try:
                report_stats = sync_report_range(client, store, report_run_id, advertiser_id, args)
                store.finish_run(report_run_id, "success", report_stats)
                summary["checked_accounts"] += 1
                spend = float(report_stats.get("花费") or 0)
                if spend <= 0:
                    continue

                summary["spend_accounts"] += 1
                summary["spenders"].append({"advertiser_id": advertiser_id, "花费": spend})
                if args.report_only:
                    continue

                sync_run_id = store.start_run("spenders_sync", advertiser_id)
                try:
                    sync_stats = sync_advertiser(client, store, sync_run_id, advertiser_id, args)
                    store.finish_run(sync_run_id, "success", sync_stats)
                    summary["synced_accounts"] += 1
                except Exception as error:
                    store.finish_run(sync_run_id, "failed", message=str(error))
                    summary["errors"] += 1
                    if not args.continue_on_error:
                        raise
            except Exception as error:
                store.finish_run(report_run_id, "failed", message=str(error))
                summary["errors"] += 1
                if not args.continue_on_error:
                    raise
        summary["spenders"] = sorted(summary["spenders"], key=lambda item: item["花费"], reverse=True)
        print_json({"ok": True, "db": str(store.db_path), "summary": summary})
    finally:
        store.close()


def command_sync_spenders(args):
    store = create_store(args)
    client = KuaishouClient()
    rows = store.conn.execute(
        """
        SELECT "广告主ID", "花费"
        FROM "快手有消耗账户"
        ORDER BY "花费" DESC
        """
    ).fetchall()
    if args.limit:
        rows = rows[: args.limit]
    summary = {"accounts": len(rows), "synced_accounts": 0, "errors": 0, "details": []}
    try:
        for index, row in enumerate(rows, start=1):
            advertiser_id = int(row["广告主ID"])
            print(f"[{index}/{len(rows)}] 同步有消耗广告主 {advertiser_id}", file=sys.stderr)
            run_id = store.start_run("sync_spender", advertiser_id)
            try:
                stats = sync_advertiser(client, store, run_id, advertiser_id, args)
                store.finish_run(run_id, "success", stats)
                summary["synced_accounts"] += 1
                summary["details"].append({"advertiser_id": advertiser_id, **stats})
            except Exception as error:
                store.finish_run(run_id, "failed", message=str(error))
                summary["errors"] += 1
                summary["details"].append({"advertiser_id": advertiser_id, "error": str(error)})
                if not args.continue_on_error:
                    raise
        print_json({"ok": True, "db": str(store.db_path), "summary": summary})
    finally:
        store.close()


def run_scheduled_pull(args, log_file):
    started_at = shanghai_now()
    today = started_at.date().isoformat()
    log_line(f"scheduled-pull started date={today}")
    excel_path = ensure_demand_excel()
    accounts = read_enabled_demand_accounts(excel_path)
    summary = {
        "date": today,
        "excel": str(excel_path),
        "log_file": str(log_file),
        "accounts": len(accounts),
        "report_success": 0,
        "config_success": 0,
        "errors": 0,
        "details": [],
    }
    if not accounts:
        log_line("no enabled demand accounts found")
        summary["finished_at"] = shanghai_now().isoformat()
        return {"ok": True, "summary": summary}

    store = create_store_with_retry(args, init_schema=False, attempts=6, delay_seconds=10)
    client = KuaishouClient()
    try:
        stale_runs = run_sqlite_write_with_retry(
            "mark stale running runs",
            lambda: store.mark_stale_running_runs(
                started_at.replace(microsecond=0).strftime("%Y-%m-%d %H:%M:%S"),
                "stale running task marked failed before next scheduled pull",
            ),
        )
        if stale_runs:
            log_line(f"marked stale running runs failed count={stale_runs}")

        try:
            token_result = client.ensure_access_token(leeway_seconds=3600)
            log_line("access token ready" + (" (refresh skipped)" if token_result.get("skipped") else " (refreshed)"))
        except Exception as error:
            log_line(f"access token refresh skipped/failed: {error}")
            if is_token_error(error):
                alert = maybe_alert_token_problem(f"刷新 token 失败：{error}")
                summary["errors"] += 1
                summary["token_error"] = str(error)
                summary["alert"] = alert
                summary["finished_at"] = shanghai_now().isoformat()
                log_line(f"token alert generated {json.dumps(alert, ensure_ascii=False)}")
                log_line(f"scheduled-pull stopped summary={json.dumps(summary, ensure_ascii=False)}")
                return {"ok": False, "summary": summary}

        for index, account in enumerate(accounts, start=1):
            advertiser_id = int(account["广告主ID"])
            detail = {
                "advertiser_id": advertiser_id,
                "advertiser_name": account.get("广告主名称", ""),
                "report": None,
                "config": None,
                "errors": [],
            }
            log_line(f"[{index}/{len(accounts)}] pull advertiser_id={advertiser_id} name={detail['advertiser_name']}")

            report_run_id = store.start_run("scheduled_report", advertiser_id)
            report_args = make_child_args(args, start_date=today, end_date=today)
            try:
                report_stats = sync_report_range(client, store, report_run_id, advertiser_id, report_args)
                store.finish_run(report_run_id, "success", report_stats)
                summary["report_success"] += 1
                detail["report"] = {"run_id": report_run_id, **report_stats}
                log_line(f"  report ok run_id={report_run_id} stats={json.dumps(report_stats, ensure_ascii=False)}")
            except Exception as error:
                store.finish_run(report_run_id, "failed", message=str(error))
                summary["errors"] += 1
                detail["errors"].append({"stage": "report", "run_id": report_run_id, "message": str(error)})
                log_line(f"  report failed run_id={report_run_id} error={error}")
                if is_token_error(error):
                    alert = maybe_alert_token_problem(f"拉取日报失败，token 异常：{error}")
                    summary["token_error"] = str(error)
                    summary["alert"] = alert
                    summary["details"].append(detail)
                    summary["finished_at"] = shanghai_now().isoformat()
                    log_line(f"token alert generated {json.dumps(alert, ensure_ascii=False)}")
                    log_line(f"scheduled-pull stopped summary={json.dumps(summary, ensure_ascii=False)}")
                    return {"ok": False, "summary": summary}

            sync_run_id = store.start_run("scheduled_sync", advertiser_id)
            try:
                sync_stats = sync_advertiser(client, store, sync_run_id, advertiser_id, args)
                store.finish_run(sync_run_id, "success", sync_stats)
                summary["config_success"] += 1
                detail["config"] = {"run_id": sync_run_id, **sync_stats}
                log_line(f"  config ok run_id={sync_run_id} stats={json.dumps(sync_stats, ensure_ascii=False)}")
            except Exception as error:
                store.finish_run(sync_run_id, "failed", message=str(error))
                summary["errors"] += 1
                detail["errors"].append({"stage": "config", "run_id": sync_run_id, "message": str(error)})
                log_line(f"  config failed run_id={sync_run_id} error={error}")
                if is_token_error(error):
                    alert = maybe_alert_token_problem(f"拉取配置失败，token 异常：{error}")
                    summary["token_error"] = str(error)
                    summary["alert"] = alert
                    summary["details"].append(detail)
                    summary["finished_at"] = shanghai_now().isoformat()
                    log_line(f"token alert generated {json.dumps(alert, ensure_ascii=False)}")
                    log_line(f"scheduled-pull stopped summary={json.dumps(summary, ensure_ascii=False)}")
                    return {"ok": False, "summary": summary}

            summary["details"].append(detail)

        if getattr(args, "keep_today_only", False) and (summary["report_success"] > 0 or summary["config_success"] > 0):
            cleanup_stats = store.keep_only_today(today)
            summary["cleanup"] = cleanup_stats
            log_line(f"cleanup keep-today-only stats={json.dumps(cleanup_stats, ensure_ascii=False)}")
        elif getattr(args, "keep_today_only", False):
            summary["cleanup_skipped"] = "no successful report/config pull; keep existing data"
            log_line("cleanup skipped because no successful report/config pull")

        automation_result = run_automation_after_pull(args)
        summary["automation"] = automation_result
        log_line(f"automation summary={json.dumps(automation_result, ensure_ascii=False)}")

        summary["finished_at"] = shanghai_now().isoformat()
        log_line(f"scheduled-pull finished summary={json.dumps(summary, ensure_ascii=False)}")
        return {"ok": summary["errors"] == 0, "summary": summary}
    finally:
        store.close()


def make_child_args(args, **overrides):
    values = vars(args).copy()
    values.update(overrides)
    return argparse.Namespace(**values)


def run_automation_after_pull(args):
    from automation_center import AutomationCenter

    center = AutomationCenter(args.db)
    try:
        return center.run_rules(source="scheduled_pull", dry_run=None, respect_enabled=True)
    except Exception as error:
        return {
            "ok": False,
            "status": "failed",
            "reason": str(error),
        }
    finally:
        center.close()


def ensure_demand_excel():
    if DEMAND_EXCEL.exists():
        workbook = load_workbook(DEMAND_EXCEL)
        sheet = workbook["需求账号"] if "需求账号" in workbook.sheetnames else workbook.create_sheet("需求账号")
        changed = ensure_demand_sheet(sheet)
        if changed:
            workbook.save(DEMAND_EXCEL)
        return DEMAND_EXCEL

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "需求账号"
    ensure_demand_sheet(sheet)
    workbook.save(DEMAND_EXCEL)
    return DEMAND_EXCEL


def ensure_demand_sheet(sheet):
    headers = ["启用", "广告主ID", "广告主名称", "备注"]
    changed = False
    existing_headers = [sheet.cell(row=1, column=index).value for index in range(1, len(headers) + 1)]
    if existing_headers != headers:
        for index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=index, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4472C4")
        changed = True

    has_account = False
    for row in range(2, sheet.max_row + 1):
        if str(sheet.cell(row=row, column=2).value or "").strip() == str(DEFAULT_DEMAND_ACCOUNT["广告主ID"]):
            has_account = True
            break
    if not has_account:
        row = max(2, sheet.max_row + 1)
        for column, header in enumerate(headers, start=1):
            sheet.cell(row=row, column=column, value=DEFAULT_DEMAND_ACCOUNT[header])
        changed = True

    widths = {"A": 10, "B": 16, "C": 24, "D": 28}
    for column, width in widths.items():
        if sheet.column_dimensions[column].width != width:
            sheet.column_dimensions[column].width = width
            changed = True
    sheet.freeze_panes = "A2"
    return changed


def read_enabled_demand_accounts(path):
    workbook = load_workbook(path, data_only=True)
    if "需求账号" not in workbook.sheetnames:
        return []
    sheet = workbook["需求账号"]
    headers = {}
    for index, cell in enumerate(sheet[1], start=1):
        value = str(cell.value or "").strip()
        if value:
            headers[value] = index
    required = ["启用", "广告主ID"]
    missing = [name for name in required if name not in headers]
    if missing:
        raise SystemExit(f"需求账号.xlsx 缺少字段: {', '.join(missing)}")

    accounts = []
    seen = set()
    for row in range(2, sheet.max_row + 1):
        enabled = str(sheet.cell(row=row, column=headers["启用"]).value or "").strip().lower()
        if enabled not in {"是", "yes", "y", "true", "1", "启用"}:
            continue
        raw_id = sheet.cell(row=row, column=headers["广告主ID"]).value
        try:
            advertiser_id = int(str(raw_id).strip())
        except (TypeError, ValueError):
            continue
        if advertiser_id in seen:
            continue
        seen.add(advertiser_id)
        name = ""
        remark = ""
        if "广告主名称" in headers:
            name = str(sheet.cell(row=row, column=headers["广告主名称"]).value or "").strip()
        if "备注" in headers:
            remark = str(sheet.cell(row=row, column=headers["备注"]).value or "").strip()
        accounts.append({"广告主ID": advertiser_id, "广告主名称": name, "备注": remark})
    return accounts


def powershell_quote(value):
    return "'" + str(value).replace("'", "''") + "'"


def sync_advertiser(client, store, run_id, advertiser_id, args, campaign_id=None):
    advertiser_id = int(advertiser_id)
    page_size = args.page_size
    stats = {"campaigns": 0, "units": 0, "creatives": 0}

    def save_raw(endpoint, request_body, response):
        store.save_raw_response(run_id, endpoint, request_body, response)

    campaign_body = {"advertiser_id": advertiser_id}
    if campaign_id:
        campaign_body["campaign_ids"] = [int(campaign_id)]
    campaigns = list_all_with_retry(client, CAMPAIGN_LIST, campaign_body, args, raw_callback=save_raw)
    stats["campaigns"] = store.upsert_campaigns(advertiser_id, campaigns["details"], run_id)

    unit_body = {"advertiser_id": advertiser_id}
    if campaign_id:
        unit_body["campaign_id"] = int(campaign_id)
    units = list_all_with_retry(client, UNIT_LIST, unit_body, args, raw_callback=save_raw)
    stats["units"] = store.upsert_units(advertiser_id, units["details"], run_id)

    creative_body = {"advertiser_id": advertiser_id}
    if campaign_id:
        creative_body["campaign_id"] = int(campaign_id)
    creatives = list_all_with_retry(client, CREATIVE_LIST, creative_body, args, raw_callback=save_raw)
    stats["creatives"] = store.upsert_creatives(advertiser_id, creatives["details"], run_id)
    return stats


def sync_report_range(client, store, run_id, advertiser_id, args):
    request_body = {
        "advertiser_id": int(advertiser_id),
        "start_date": args.start_date,
        "end_date": args.end_date,
    }
    if getattr(args, "stat_hour", None) is not None:
        request_body["stat_hour"] = int(args.stat_hour)

    def save_raw(endpoint, body, response):
        store.save_raw_response(run_id, endpoint, body, response)

    result = list_all_with_retry(client, ACCOUNT_REPORT, request_body, args, raw_callback=save_raw)
    count = store.upsert_daily_reports(
        advertiser_id,
        result["details"],
        run_id,
        source="快手接口",
        source_file="",
    )
    spend = 0.0
    for row in result["details"]:
        try:
            spend += float(row.get("charge") or row.get("花费") or 0)
        except (TypeError, ValueError):
            pass

    unit_result = list_all_with_retry(client, UNIT_REPORT, request_body, args, raw_callback=save_raw)
    unit_count = store.upsert_unit_daily_reports(
        advertiser_id,
        unit_result["details"],
        run_id,
        source="快手接口",
    )
    unit_spend = 0.0
    for row in unit_result["details"]:
        try:
            unit_spend += float(row.get("charge") or row.get("花费") or 0)
        except (TypeError, ValueError):
            pass
    return {"快手日报": count, "花费": spend, "快手广告组日报": unit_count, "广告组花费": unit_spend}


def list_all_with_retry(client, path, body, args, raw_callback=None):
    attempts = max(1, int(getattr(args, "retries", 3) or 3))
    delay = float(getattr(args, "retry_delay", 5.0) or 5.0)
    last_error = None
    for attempt in range(1, attempts + 1):
        try:
            return client.list_all(path, body, page_size=args.page_size, raw_callback=raw_callback)
        except KuaishouApiError as error:
            last_error = error
            if not is_rate_limit_error(error) or attempt >= attempts:
                raise
            wait_seconds = delay * attempt
            print(f"  触发限流，等待 {wait_seconds:.0f}s 后重试 ({attempt}/{attempts})", file=sys.stderr)
            time.sleep(wait_seconds)
    raise last_error


def is_rate_limit_error(error):
    text = str(error)
    body = getattr(error, "body", None)
    if isinstance(body, dict):
        text = f"{text} {body.get('message', '')} {body.get('code', '')}"
    return "限流" in text or "rate" in text.lower()


def resolve_advertiser_ids(args):
    values = list(args.advertiser_id or [])
    if not values:
        env_value = os.getenv("KUAISHOU_ADVERTISER_ID", "")
        values = [item.strip() for item in env_value.replace("，", ",").split(",") if item.strip()]
    return [int(value) for value in values if str(value).strip()]


def build_parser():
    parser = argparse.ArgumentParser(description="快手 OpenAPI 数据落库工具")
    parser.add_argument("--db", help="SQLite 数据库路径，默认读取 KS_DB_PATH 或 data/kuaishou.db")
    parser.add_argument("--page-size", type=int, default=int(os.getenv("KS_PAGE_SIZE", "100") or 100))
    subparsers = parser.add_subparsers(dest="command", required=True)

    init_db = subparsers.add_parser("init-db", help="初始化 SQLite 表结构")
    init_db.set_defaults(func=command_init_db)

    status = subparsers.add_parser("status", help="查看当前鉴权配置状态")
    status.set_defaults(func=command_status)

    exchange = subparsers.add_parser("exchange-token", help="用 auth_code 换取 access_token")
    exchange.add_argument("auth_code")
    exchange.set_defaults(func=command_exchange)

    refresh = subparsers.add_parser("refresh-token", help="刷新 access_token")
    refresh.set_defaults(func=command_refresh)

    check_token = subparsers.add_parser("check-token", help="检测 token 是否可刷新，失败时写入告警")
    check_token.set_defaults(func=command_check_token)

    accounts = subparsers.add_parser("accounts", help="同步已授权广告账户列表")
    accounts.add_argument("--auth-user-id", help="授权主体 ID，默认读取 KUAISHOU_AUTH_USER_ID")
    accounts.set_defaults(func=command_accounts)

    sync = subparsers.add_parser("sync", help="同步广告计划、广告组、创意")
    sync.add_argument("--advertiser-id", action="append", help="广告主 ID，可重复传多个")
    sync.add_argument("--continue-on-error", action="store_true", help="多账户同步时某个账户失败后继续")
    sync.set_defaults(func=command_sync)

    snapshot = subparsers.add_parser("snapshot", help="只同步某个计划及其广告组/创意")
    snapshot.add_argument("--advertiser-id", required=True)
    snapshot.add_argument("--campaign-id", required=True)
    snapshot.set_defaults(func=command_snapshot)

    report = subparsers.add_parser("report", help="同步广告主日报，并写入中文表 快手日报")
    report.add_argument("--advertiser-id", action="append", help="广告主 ID，可重复传多个")
    report.add_argument("--start-date", required=True, help="开始日期，格式 yyyy-MM-dd")
    report.add_argument("--end-date", required=True, help="结束日期，格式 yyyy-MM-dd")
    report.add_argument("--stat-hour", type=int, help="可选：同步指定小时")
    report.add_argument("--continue-on-error", action="store_true", help="多账户同步时某个账户失败后继续")
    report.set_defaults(func=command_report)

    report_csv = subparsers.add_parser("import-report-csv", help="导入快手 CSV 报表，并写入中文表 快手日报")
    report_csv.add_argument("--advertiser-id", required=True)
    report_csv.add_argument("--file", required=True)
    report_csv.set_defaults(func=command_import_report_csv)

    spenders = subparsers.add_parser("spenders", help="抓取所有有消耗账户的数据")
    spenders.add_argument("--start-date", required=True, help="开始日期，格式 yyyy-MM-dd")
    spenders.add_argument("--end-date", required=True, help="结束日期，格式 yyyy-MM-dd")
    spenders.add_argument("--stat-hour", type=int, help="可选：同步指定小时")
    spenders.add_argument("--limit", type=int, help="调试用：只检查前 N 个账户")
    spenders.add_argument("--report-only", action="store_true", help="只抓日报并筛选有消耗账户，不同步计划/广告组/创意")
    spenders.add_argument("--continue-on-error", action="store_true", default=True, help="某个账户失败后继续")
    spenders.add_argument("--request-delay", type=float, default=3.0, help="账户之间等待秒数，默认 3 秒")
    spenders.add_argument("--retries", type=int, default=6, help="限流自动重试次数，默认 6")
    spenders.add_argument("--retry-delay", type=float, default=20.0, help="限流首次等待秒数，后续线性递增，默认 20 秒")
    spenders.set_defaults(func=command_spenders)

    sync_spenders = subparsers.add_parser("sync-spenders", help="同步 快手有消耗账户 里的计划/广告组/创意")
    sync_spenders.add_argument("--limit", type=int, help="调试用：只同步前 N 个账户")
    sync_spenders.add_argument("--continue-on-error", action="store_true", default=True, help="某个账户失败后继续")
    sync_spenders.set_defaults(func=command_sync_spenders)

    init_demand_excel = subparsers.add_parser("init-demand-excel", help="创建或补齐项目内 需求账号.xlsx")
    init_demand_excel.set_defaults(func=command_init_demand_excel)

    automation_status = subparsers.add_parser("automation-status", help="查看自动投放中台状态")
    automation_status.set_defaults(func=command_automation_status)

    automation_run = subparsers.add_parser("automation-run", help="手动执行一次自动投放规则")
    automation_run.add_argument("--execute", action="store_true", help="真实调用快手接口；不传则只试运行")
    automation_run.add_argument("--ignore-disabled", action="store_true", help="忽略自动执行开关，手动强制跑规则")
    automation_run.set_defaults(func=command_automation_run)

    web = subparsers.add_parser("web", help="启动本机 Web 自动投放中台")
    web.add_argument("--host", default="127.0.0.1", help="监听地址，默认 127.0.0.1")
    web.add_argument("--port", type=int, default=8787, help="监听端口，默认 8787")
    web.set_defaults(func=command_web)

    scheduled_pull = subparsers.add_parser("scheduled-pull", help="读取 需求账号.xlsx，每 30 分钟拉今天累计数据")
    scheduled_pull.add_argument("--retries", type=int, default=6, help="限流自动重试次数，默认 6")
    scheduled_pull.add_argument("--retry-delay", type=float, default=20.0, help="限流首次等待秒数，后续线性递增，默认 20 秒")
    scheduled_pull.add_argument("--keep-today-only", action=argparse.BooleanOptionalAction, default=False, help="只保留当天最新入库数据，默认关闭以支持历史回看")
    scheduled_pull.set_defaults(func=command_scheduled_pull)

    install_schedule = subparsers.add_parser("install-schedule", help="创建 Windows 计划任务，每 30 分钟执行 scheduled-pull")
    install_schedule.set_defaults(func=command_install_schedule)

    return parser


def main(argv=None):
    load_runtime_env()
    parser = build_parser()
    args = parser.parse_args(argv)
    try:
        args.func(args)
    except KuaishouApiError as error:
        print_json({"ok": False, "error": str(error), "status": error.status, "detail": error.body})
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
