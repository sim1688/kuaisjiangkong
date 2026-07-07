import json
import math
import time
import uuid
from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import load_workbook

from kuaishou_client import KuaishouApiError, KuaishouClient
from kuaishou_store import KuaishouStore, local_now


ROOT_DIR = Path(__file__).resolve().parent
DEMAND_EXCEL = ROOT_DIR / "需求账号.xlsx"
SHANGHAI_TZ = timezone(timedelta(hours=8), name="Asia/Shanghai")

UNIT_STATUS_ENDPOINT = "/rest/openapi/v1/ad_unit/update/status"
UNIT_UPDATE_ENDPOINT = "/rest/openapi/gw/dsp/unit/update"
UNIT_BUDGET_ENDPOINT = "/rest/openapi/v1/ad_unit/update/day_budget"
ADVERTISER_FUND_ENDPOINT = "/rest/openapi/v1/advertiser/fund/get"

DEFAULT_SETTINGS = {
    "auto_enabled": False,
    "dry_run": True,
    "max_actions_per_run": 20,
    "max_actions_per_account_per_run": 5,
    "max_actions_per_unit_per_day": 1,
    "request_delay_seconds": 0.5,
    "blocked_advertiser_ids": [],
    "blocked_campaign_ids": [],
    "blocked_unit_ids": [],
}

FIELD_DEFS = [
    {"key": "advertiser_id", "label": "广告主ID", "type": "number"},
    {"key": "campaign_id", "label": "计划ID", "type": "number"},
    {"key": "campaign_name", "label": "计划名称", "type": "text"},
    {"key": "campaign_put_status", "label": "计划投放状态", "type": "number"},
    {"key": "unit_id", "label": "广告组ID", "type": "number"},
    {"key": "unit_name", "label": "广告组名称", "type": "text"},
    {"key": "account_balance", "label": "账号余额", "type": "number"},
    {"key": "spend", "label": "今日花费", "type": "number"},
    {"key": "roi_iaa", "label": "IAA广告变现ROI", "type": "number"},
    {"key": "yesterday_spend", "label": "昨天花费", "type": "number"},
    {"key": "yesterday_roi_iaa", "label": "昨天IAA广告变现ROI", "type": "number"},
    {"key": "roi_day", "label": "当日广告变现ROI", "type": "number"},
    {"key": "roi_week", "label": "激活七日ROI", "type": "number"},
    {"key": "roi_ratio", "label": "当前ROI系数", "type": "number"},
    {"key": "put_status", "label": "广告组投放状态", "type": "number"},
    {"key": "status", "label": "广告组平台状态", "type": "number"},
    {"key": "day_budget_yuan", "label": "广告组日预算(元)", "type": "number"},
]

ACTION_DEFS = [
    {"key": "pause_unit", "label": "暂停广告组", "needs_value": False},
    {"key": "enable_unit", "label": "开启广告组", "needs_value": False},
    {"key": "set_roi_ratio", "label": "修改ROI系数", "needs_value": True, "value_label": "ROI系数"},
    {"key": "set_day_budget", "label": "修改广告组预算", "needs_value": True, "value_label": "预算(元)"},
]

VALUE_ACTIONS = {"set_roi_ratio", "set_day_budget"}
ACTION_MODES = {"set", "add_value", "add_percent"}


def shanghai_today():
    return datetime.now(SHANGHAI_TZ).date().isoformat()


def shanghai_yesterday():
    return (datetime.now(SHANGHAI_TZ).date() - timedelta(days=1)).isoformat()


def normalize_report_date(value=None):
    if value in (None, ""):
        return shanghai_today()
    text = str(value).strip()
    try:
        return datetime.strptime(text, "%Y-%m-%d").date().isoformat()
    except ValueError:
        return shanghai_today()


def previous_report_date(value):
    day = datetime.strptime(normalize_report_date(value), "%Y-%m-%d").date()
    return (day - timedelta(days=1)).isoformat()


def json_text(value):
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def json_loads(value, default=None):
    if value in (None, ""):
        return default
    try:
        return json.loads(value)
    except (TypeError, json.JSONDecodeError):
        return default


def to_number(value):
    if value in (None, ""):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if math.isnan(number) or math.isinf(number):
        return None
    return number


def normalize_id_list(value):
    if value in (None, ""):
        return []
    if isinstance(value, str):
        items = value.replace("，", ",").split(",")
    elif isinstance(value, (list, tuple, set)):
        items = value
    else:
        items = [value]
    result = []
    for item in items:
        text = str(item).strip()
        if not text:
            continue
        try:
            result.append(int(float(text)))
        except ValueError:
            continue
    return sorted(set(result))


def money_yuan_to_li(value):
    number = to_number(value)
    if number is None:
        raise ValueError("预算必须是数字")
    if number < 0:
        raise ValueError("预算不能小于 0")
    return int(round(number * 1000))


def li_to_yuan(value):
    number = to_number(value)
    if number is None:
        return None
    return round(number / 1000, 3)


def read_enabled_demand_account_ids(path=DEMAND_EXCEL):
    if not path.exists():
        return []
    workbook = load_workbook(path, data_only=True)
    if "需求账号" not in workbook.sheetnames:
        return []
    sheet = workbook["需求账号"]
    headers = {}
    for index, cell in enumerate(sheet[1], start=1):
        value = str(cell.value or "").strip()
        if value:
            headers[value] = index
    if "启用" not in headers or "广告主ID" not in headers:
        return []

    account_ids = []
    seen = set()
    for row in range(2, sheet.max_row + 1):
        enabled = str(sheet.cell(row=row, column=headers["启用"]).value or "").strip().lower()
        if enabled not in {"是", "yes", "y", "true", "1", "启用"}:
            continue
        raw_id = sheet.cell(row=row, column=headers["广告主ID"]).value
        try:
            advertiser_id = int(str(raw_id).strip())
        except (TypeError, ValueError):
            continue
        if advertiser_id not in seen:
            account_ids.append(advertiser_id)
            seen.add(advertiser_id)
    return account_ids


class ActionSkipped(RuntimeError):
    def __init__(self, message, request=None):
        super().__init__(message)
        self.request = request or {}


class AutomationCenter:
    def __init__(self, db_path=None, client=None, init_schema=True, db_timeout=60, busy_timeout_ms=60000):
        self.store = KuaishouStore(db_path, timeout=db_timeout, busy_timeout_ms=busy_timeout_ms, set_wal=init_schema)
        if init_schema:
            self.store.init_schema()
        self.conn = self.store.conn
        self.client = client or KuaishouClient()
        if init_schema:
            self.init_schema()

    def close(self):
        self.store.close()

    def init_schema(self):
        self.conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS automation_settings (
              key TEXT PRIMARY KEY,
              value_json TEXT NOT NULL,
              updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS automation_rules (
              rule_id INTEGER PRIMARY KEY AUTOINCREMENT,
              name TEXT NOT NULL,
              enabled INTEGER NOT NULL DEFAULT 1,
              priority INTEGER NOT NULL DEFAULT 100,
              max_daily_triggers INTEGER NOT NULL DEFAULT 1,
              match_type TEXT NOT NULL DEFAULT 'all',
              conditions_json TEXT NOT NULL DEFAULT '[]',
              action_json TEXT NOT NULL DEFAULT '{}',
              created_at TEXT NOT NULL,
              updated_at TEXT NOT NULL
            );

            CREATE INDEX IF NOT EXISTS idx_automation_rules_enabled_priority
            ON automation_rules(enabled, priority, rule_id);

            CREATE TABLE IF NOT EXISTS automation_logs (
              log_id INTEGER PRIMARY KEY AUTOINCREMENT,
              batch_id TEXT NOT NULL,
              source TEXT NOT NULL,
              rule_id INTEGER,
              rule_name TEXT,
              advertiser_id INTEGER,
              campaign_id INTEGER,
              campaign_name TEXT,
              unit_id INTEGER,
              unit_name TEXT,
              action_type TEXT,
              action_value_json TEXT NOT NULL DEFAULT '{}',
              status TEXT NOT NULL,
              reason TEXT,
              request_json TEXT NOT NULL DEFAULT '{}',
              response_json TEXT NOT NULL DEFAULT '{}',
              before_json TEXT NOT NULL DEFAULT '{}',
              created_at TEXT NOT NULL
            );

            CREATE INDEX IF NOT EXISTS idx_automation_logs_created_at
            ON automation_logs(created_at);

            CREATE INDEX IF NOT EXISTS idx_automation_logs_unit_day
            ON automation_logs(unit_id, created_at, status);

            CREATE TABLE IF NOT EXISTS advertiser_funds (
              advertiser_id INTEGER PRIMARY KEY,
              balance REAL,
              recharge_balance REAL,
              contract_rebate REAL,
              direct_rebate REAL,
              raw_json TEXT NOT NULL DEFAULT '{}',
              updated_at TEXT NOT NULL
            );
            """
        )
        self._ensure_column("automation_rules", "max_daily_triggers", "INTEGER NOT NULL DEFAULT 1")
        now = local_now()
        for key, value in DEFAULT_SETTINGS.items():
            self.conn.execute(
                """
                INSERT OR IGNORE INTO automation_settings(key, value_json, updated_at)
                VALUES (?, ?, ?)
                """,
                (key, json_text(value), now),
            )
        self.conn.commit()

    def _ensure_column(self, table, column, definition):
        columns = {row["name"] for row in self.conn.execute(f"PRAGMA table_info({table})").fetchall()}
        if column not in columns:
            self.conn.execute(f"ALTER TABLE {table} ADD COLUMN {column} {definition}")

    def get_settings(self):
        settings = dict(DEFAULT_SETTINGS)
        rows = self.conn.execute("SELECT key, value_json FROM automation_settings").fetchall()
        for row in rows:
            if row["key"] in settings:
                settings[row["key"]] = json_loads(row["value_json"], settings[row["key"]])
        for key in ("blocked_advertiser_ids", "blocked_campaign_ids", "blocked_unit_ids"):
            settings[key] = normalize_id_list(settings.get(key))
        settings["auto_enabled"] = bool(settings.get("auto_enabled"))
        settings["dry_run"] = bool(settings.get("dry_run"))
        settings["max_actions_per_run"] = max(0, int(settings.get("max_actions_per_run") or 0))
        settings["max_actions_per_account_per_run"] = max(
            0, int(settings.get("max_actions_per_account_per_run") or 0)
        )
        settings["max_actions_per_unit_per_day"] = max(0, int(settings.get("max_actions_per_unit_per_day") or 0))
        settings["request_delay_seconds"] = max(0.0, float(settings.get("request_delay_seconds") or 0))
        return settings

    def update_settings(self, values):
        current = self.get_settings()
        next_values = dict(current)
        for key in DEFAULT_SETTINGS:
            if key not in values:
                continue
            if key in {"auto_enabled", "dry_run"}:
                next_values[key] = bool(values[key])
            elif key in {"blocked_advertiser_ids", "blocked_campaign_ids", "blocked_unit_ids"}:
                next_values[key] = normalize_id_list(values[key])
            elif key == "request_delay_seconds":
                next_values[key] = max(0.0, float(values[key] or 0))
            else:
                next_values[key] = max(0, int(values[key] or 0))

        now = local_now()
        for key, value in next_values.items():
            self.conn.execute(
                """
                INSERT INTO automation_settings(key, value_json, updated_at)
                VALUES (?, ?, ?)
                ON CONFLICT(key) DO UPDATE SET
                  value_json = excluded.value_json,
                  updated_at = excluded.updated_at
                """,
                (key, json_text(value), now),
            )
        self.conn.commit()
        return self.get_settings()

    def list_rules(self, enabled_only=False):
        sql = "SELECT * FROM automation_rules"
        params = []
        if enabled_only:
            sql += " WHERE enabled = 1"
        sql += " ORDER BY priority ASC, rule_id ASC"
        rows = self.conn.execute(sql, params).fetchall()
        return [self._rule_from_row(row) for row in rows]

    def save_rule(self, payload):
        name = str(payload.get("name") or "").strip()
        if not name:
            raise ValueError("规则名称不能为空")
        priority = int(payload.get("priority") or 100)
        max_daily_triggers = max(0, int(payload.get("max_daily_triggers") if payload.get("max_daily_triggers") not in (None, "") else 1))
        enabled = 1 if payload.get("enabled", True) else 0
        match_type = str(payload.get("match_type") or "all").lower()
        if match_type not in {"all", "any"}:
            match_type = "all"
        conditions = self._normalize_conditions(payload.get("conditions") or [])
        action = self._normalize_action(payload.get("action") or {})
        now = local_now()
        rule_id = payload.get("rule_id")
        if rule_id:
            self.conn.execute(
                """
                UPDATE automation_rules
                SET name = ?, enabled = ?, priority = ?, max_daily_triggers = ?, match_type = ?,
                    conditions_json = ?, action_json = ?, updated_at = ?
                WHERE rule_id = ?
                """,
                (
                    name,
                    enabled,
                    priority,
                    max_daily_triggers,
                    match_type,
                    json_text(conditions),
                    json_text(action),
                    now,
                    int(rule_id),
                ),
            )
        else:
            cursor = self.conn.execute(
                """
                INSERT INTO automation_rules(
                  name, enabled, priority, max_daily_triggers, match_type,
                  conditions_json, action_json, created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    name,
                    enabled,
                    priority,
                    max_daily_triggers,
                    match_type,
                    json_text(conditions),
                    json_text(action),
                    now,
                    now,
                ),
            )
            rule_id = cursor.lastrowid
        self.conn.commit()
        return self.get_rule(int(rule_id))

    def get_rule(self, rule_id):
        row = self.conn.execute("SELECT * FROM automation_rules WHERE rule_id = ?", (int(rule_id),)).fetchone()
        if not row:
            raise KeyError(f"规则不存在: {rule_id}")
        return self._rule_from_row(row)

    def delete_rule(self, rule_id):
        cursor = self.conn.execute("DELETE FROM automation_rules WHERE rule_id = ?", (int(rule_id),))
        self.conn.commit()
        return cursor.rowcount

    def list_logs(self, limit=100):
        limit = max(1, min(500, int(limit or 100)))
        rows = self.conn.execute(
            """
            SELECT *
            FROM automation_logs
            ORDER BY log_id DESC
            LIMIT ?
            """,
            (limit,),
        ).fetchall()
        return [self._log_from_row(row) for row in rows]

    def data_version(self):
        row = self.conn.execute(
            """
            SELECT
              (SELECT COALESCE(MAX("最后入库时间"), '') FROM "快手广告组日报") AS unit_report_at,
              (SELECT COALESCE(MAX(last_seen_at), '') FROM units) AS units_at,
              (SELECT COALESCE(MAX(log_id), 0) FROM automation_logs) AS log_id,
              (SELECT COALESCE(MAX(updated_at), '') FROM automation_rules) AS rules_at,
              (SELECT COALESCE(MAX(updated_at), '') FROM automation_settings) AS settings_at
            """
        ).fetchone()
        value = dict(row)
        value["version"] = "|".join(str(value[key]) for key in sorted(value))
        return value

    def refresh_account_funds(self, advertiser_ids, max_age_seconds=300):
        refreshed = []
        errors = []
        now_dt = datetime.now(SHANGHAI_TZ).replace(microsecond=0)
        now_text = now_dt.strftime("%Y-%m-%d %H:%M:%S")
        for advertiser_id in normalize_id_list(advertiser_ids):
            row = self.conn.execute(
                "SELECT updated_at FROM advertiser_funds WHERE advertiser_id = ?",
                (advertiser_id,),
            ).fetchone()
            if row:
                try:
                    updated_at = datetime.strptime(row["updated_at"], "%Y-%m-%d %H:%M:%S").replace(tzinfo=SHANGHAI_TZ)
                    if (now_dt - updated_at).total_seconds() < max_age_seconds:
                        continue
                except ValueError:
                    pass
            try:
                response = self.client.request(
                    ADVERTISER_FUND_ENDPOINT,
                    method="POST",
                    body={"advertiser_id": advertiser_id},
                )
                self.client._assert_success(response)
                data = response.get("data") if isinstance(response, dict) else {}
                if not isinstance(data, dict):
                    data = {}
                self.conn.execute(
                    """
                    INSERT INTO advertiser_funds(
                      advertiser_id, balance, recharge_balance, contract_rebate,
                      direct_rebate, raw_json, updated_at
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(advertiser_id) DO UPDATE SET
                      balance = excluded.balance,
                      recharge_balance = excluded.recharge_balance,
                      contract_rebate = excluded.contract_rebate,
                      direct_rebate = excluded.direct_rebate,
                      raw_json = excluded.raw_json,
                      updated_at = excluded.updated_at
                    """,
                    (
                        advertiser_id,
                        to_number(data.get("balance")),
                        to_number(data.get("recharge_balance")),
                        to_number(data.get("contract_rebate")),
                        to_number(data.get("direct_rebate")),
                        json_text(response),
                        now_text,
                    ),
                )
                refreshed.append(advertiser_id)
            except Exception as error:
                errors.append({"advertiser_id": advertiser_id, "error": str(error)})
        self.conn.commit()
        return {"refreshed": refreshed, "errors": errors}

    def list_report_dates(self, limit=60):
        rows = self.conn.execute(
            """
            SELECT "日期" AS report_date, COUNT(*) AS rows
            FROM "快手广告组日报"
            GROUP BY "日期"
            ORDER BY "日期" DESC
            LIMIT ?
            """,
            (max(1, min(365, int(limit or 60))),),
        ).fetchall()
        return [{"date": row["report_date"], "rows": row["rows"]} for row in rows]

    def list_candidates(self, limit=200, advertiser_ids=None, report_date=None):
        ids = advertiser_ids if advertiser_ids is not None else read_enabled_demand_account_ids()
        ids = normalize_id_list(ids)
        if not ids:
            return []
        limit = max(1, min(1000, int(limit or 200)))
        selected_date = normalize_report_date(report_date)
        previous_date = previous_report_date(selected_date)
        placeholders = ",".join("?" for _ in ids)
        sql = f"""
        WITH report AS (
          SELECT
            "广告主ID" AS advertiser_id,
            "广告组ID" AS unit_id,
            SUM(COALESCE("花费", 0)) AS spend,
            CASE
              WHEN SUM(COALESCE("花费", 0)) > 0
              THEN ROUND(SUM(COALESCE("当日广告变现ROI", 0) * COALESCE("花费", 0)) / SUM(COALESCE("花费", 0)), 3)
              ELSE ROUND(AVG("当日广告变现ROI"), 3)
            END AS roi_day,
            CASE
              WHEN SUM(COALESCE("花费", 0)) > 0
              THEN ROUND(SUM(COALESCE("激活七日广告变现ROI", 0) * COALESCE("花费", 0)) / SUM(COALESCE("花费", 0)), 3)
              ELSE ROUND(AVG("激活七日广告变现ROI"), 3)
            END AS roi_week,
            CASE
              WHEN SUM(COALESCE("花费", 0)) > 0
              THEN ROUND(SUM(COALESCE("IAA广告变现ROI", 0) * COALESCE("花费", 0)) / SUM(COALESCE("花费", 0)), 3)
              ELSE ROUND(AVG("IAA广告变现ROI"), 3)
            END AS roi_iaa,
            MAX("最后入库时间") AS report_last_seen_at
          FROM "快手广告组日报"
          WHERE "日期" = ?
          GROUP BY "广告主ID", "广告组ID"
        ),
        yesterday_report AS (
          SELECT
            "广告主ID" AS advertiser_id,
            "广告组ID" AS unit_id,
            SUM(COALESCE("花费", 0)) AS yesterday_spend,
            CASE
              WHEN SUM(COALESCE("花费", 0)) > 0
              THEN ROUND(SUM(COALESCE("IAA广告变现ROI", 0) * COALESCE("花费", 0)) / SUM(COALESCE("花费", 0)), 3)
              ELSE ROUND(AVG("IAA广告变现ROI"), 3)
            END AS yesterday_roi_iaa
          FROM "快手广告组日报"
          WHERE "日期" = ?
          GROUP BY "广告主ID", "广告组ID"
        ),
        unit_history_key AS (
          SELECT advertiser_id, entity_id AS unit_id, MAX(history_id) AS history_id
          FROM entity_history
          WHERE entity_type = 'unit'
            AND substr(captured_at, 1, 10) = ?
          GROUP BY advertiser_id, entity_id
        ),
        campaign_history_key AS (
          SELECT advertiser_id, entity_id AS campaign_id, MAX(history_id) AS history_id
          FROM entity_history
          WHERE entity_type = 'campaign'
            AND substr(captured_at, 1, 10) = ?
          GROUP BY advertiser_id, entity_id
        )
        SELECT
          u.advertiser_id,
          COALESCE(json_extract(uh.raw_json, '$.campaign_id'), u.campaign_id) AS campaign_id,
          COALESCE(json_extract(ch.raw_json, '$.campaign_name'), c.campaign_name) AS campaign_name,
          COALESCE(json_extract(ch.raw_json, '$.put_status'), c.put_status) AS campaign_put_status,
          u.unit_id,
          COALESCE(json_extract(uh.raw_json, '$.unit_name'), u.unit_name) AS unit_name,
          COALESCE(json_extract(uh.raw_json, '$.roi_ratio'), u.roi_ratio) AS roi_ratio,
          COALESCE(json_extract(uh.raw_json, '$.put_status'), u.put_status) AS put_status,
          COALESCE(json_extract(uh.raw_json, '$.status'), u.status) AS status,
          COALESCE(json_extract(uh.raw_json, '$.day_budget'), u.day_budget) AS day_budget,
          COALESCE(report.report_last_seen_at, u.last_seen_at) AS last_seen_at,
          funds.balance AS account_balance,
          COALESCE(report.spend, 0) AS spend,
          report.roi_day,
          report.roi_week,
          report.roi_iaa,
          COALESCE(yesterday_report.yesterday_spend, 0) AS yesterday_spend,
          yesterday_report.yesterday_roi_iaa
        FROM units u
        LEFT JOIN unit_history_key uhk
          ON uhk.advertiser_id = u.advertiser_id
         AND uhk.unit_id = u.unit_id
        LEFT JOIN entity_history uh
          ON uh.history_id = uhk.history_id
        LEFT JOIN campaigns c
          ON c.advertiser_id = u.advertiser_id
         AND c.campaign_id = COALESCE(json_extract(uh.raw_json, '$.campaign_id'), u.campaign_id)
        LEFT JOIN campaign_history_key chk
          ON chk.advertiser_id = u.advertiser_id
         AND chk.campaign_id = COALESCE(json_extract(uh.raw_json, '$.campaign_id'), u.campaign_id)
        LEFT JOIN entity_history ch
          ON ch.history_id = chk.history_id
        LEFT JOIN report
          ON report.advertiser_id = u.advertiser_id
         AND report.unit_id = u.unit_id
        LEFT JOIN yesterday_report
          ON yesterday_report.advertiser_id = u.advertiser_id
         AND yesterday_report.unit_id = u.unit_id
        LEFT JOIN advertiser_funds funds
          ON funds.advertiser_id = u.advertiser_id
        WHERE u.advertiser_id IN ({placeholders})
        ORDER BY COALESCE(report.spend, 0) DESC, COALESCE(report.report_last_seen_at, u.last_seen_at) DESC
        LIMIT ?
        """
        rows = self.conn.execute(sql, [selected_date, previous_date, selected_date, selected_date, *ids, limit]).fetchall()
        return [self._candidate_from_row(row) for row in rows]

    def run_rules(self, source="manual", dry_run=None, respect_enabled=True):
        settings = self.get_settings()
        allowed_account_ids = set(read_enabled_demand_account_ids())
        if not allowed_account_ids:
            return {
                "ok": True,
                "status": "skipped",
                "reason": "需求账号.xlsx 没有启用账号",
                "source": source,
                "dry_run": True,
                "batch_id": "",
                "matched": 0,
                "success": 0,
                "failed": 0,
                "skipped": 0,
                "logs": [],
            }
        if respect_enabled and not settings["auto_enabled"]:
            return {
                "ok": True,
                "status": "skipped",
                "reason": "自动执行未开启",
                "source": source,
                "dry_run": True,
                "batch_id": "",
                "matched": 0,
                "success": 0,
                "failed": 0,
                "skipped": 0,
                "logs": [],
            }

        effective_dry_run = settings["dry_run"] if dry_run is None else bool(dry_run)
        batch_id = uuid.uuid4().hex[:16]
        candidates = self.list_candidates(limit=1000, advertiser_ids=allowed_account_ids)
        rules = self.list_rules(enabled_only=True)
        run_counts = {"total": 0, "success": 0, "failed": 0, "dry_run": 0, "skipped": 0, "matched": 0}
        account_counts = {}
        created_logs = []

        for candidate in candidates:
            if int(candidate["advertiser_id"]) not in allowed_account_ids:
                continue
            rule = self._first_matching_rule(candidate, rules)
            if not rule:
                continue
            run_counts["matched"] += 1
            safety = self._safety_check(rule, candidate, settings, run_counts, account_counts, allowed_account_ids)
            if safety:
                log_id = self._insert_log(
                    batch_id=batch_id,
                    source=source,
                    rule=rule,
                    candidate=candidate,
                    status="skipped",
                    reason=safety,
                    request={},
                    response={},
                )
                run_counts["skipped"] += 1
                created_logs.append(log_id)
                continue

            result = self._apply_action(rule["action"], candidate, effective_dry_run)
            log_id = self._insert_log(
                batch_id=batch_id,
                source=source,
                rule=rule,
                candidate=candidate,
                status=result["status"],
                reason=result.get("reason", ""),
                request=result.get("request", {}),
                response=result.get("response", {}),
            )
            created_logs.append(log_id)
            if result["status"] in {"success", "dry_run"}:
                run_counts["total"] += 1
                account_counts[candidate["advertiser_id"]] = account_counts.get(candidate["advertiser_id"], 0) + 1
                if result["status"] == "success":
                    run_counts["success"] += 1
                else:
                    run_counts["dry_run"] += 1
            elif result["status"] == "failed":
                run_counts["failed"] += 1
            else:
                run_counts["skipped"] += 1

            if settings["request_delay_seconds"] > 0 and not effective_dry_run:
                time.sleep(settings["request_delay_seconds"])

        logs = self.logs_by_ids(created_logs[-100:])
        return {
            "ok": run_counts["failed"] == 0,
            "status": "finished",
            "source": source,
            "dry_run": effective_dry_run,
            "batch_id": batch_id,
            **run_counts,
            "logs": logs,
        }

    def logs_by_ids(self, log_ids):
        if not log_ids:
            return []
        placeholders = ",".join("?" for _ in log_ids)
        rows = self.conn.execute(
            f"SELECT * FROM automation_logs WHERE log_id IN ({placeholders}) ORDER BY log_id DESC",
            [int(item) for item in log_ids],
        ).fetchall()
        return [self._log_from_row(row) for row in rows]

    def dashboard(self, report_date=None):
        account_ids = read_enabled_demand_account_ids()
        selected_date = normalize_report_date(report_date)
        fund_status = self.refresh_account_funds(account_ids, max_age_seconds=0)
        settings = self.get_settings()
        rules = self.list_rules()
        candidates = self.list_candidates(limit=200, advertiser_ids=account_ids, report_date=selected_date)
        logs = self.list_logs(limit=100)
        return {
            "settings": settings,
            "rules": rules,
            "logs": logs,
            "candidates": candidates,
            "report_date": selected_date,
            "previous_report_date": previous_report_date(selected_date),
            "available_report_dates": self.list_report_dates(),
            "field_defs": FIELD_DEFS,
            "action_defs": ACTION_DEFS,
            "demand_account_ids": account_ids,
            "fund_status": fund_status,
            "summary": {
                "enabled_accounts": len(account_ids),
                "rules": len(rules),
                "enabled_rules": sum(1 for rule in rules if rule["enabled"]),
                "candidates": len(candidates),
            },
        }

    def _rule_from_row(self, row):
        return {
            "rule_id": row["rule_id"],
            "name": row["name"],
            "enabled": bool(row["enabled"]),
            "priority": row["priority"],
            "max_daily_triggers": int(row["max_daily_triggers"] if row["max_daily_triggers"] is not None else 1),
            "match_type": row["match_type"],
            "conditions": json_loads(row["conditions_json"], []),
            "action": json_loads(row["action_json"], {}),
            "created_at": row["created_at"],
            "updated_at": row["updated_at"],
        }

    def _log_from_row(self, row):
        return {
            "log_id": row["log_id"],
            "batch_id": row["batch_id"],
            "source": row["source"],
            "rule_id": row["rule_id"],
            "rule_name": row["rule_name"],
            "advertiser_id": row["advertiser_id"],
            "campaign_id": row["campaign_id"],
            "campaign_name": row["campaign_name"],
            "unit_id": row["unit_id"],
            "unit_name": row["unit_name"],
            "action_type": row["action_type"],
            "action_value": json_loads(row["action_value_json"], {}),
            "status": row["status"],
            "reason": row["reason"],
            "request": json_loads(row["request_json"], {}),
            "response": json_loads(row["response_json"], {}),
            "before": json_loads(row["before_json"], {}),
            "created_at": row["created_at"],
        }

    def _candidate_from_row(self, row):
        item = dict(row)
        item["day_budget_yuan"] = li_to_yuan(item.get("day_budget"))
        item["spend"] = float(item.get("spend") or 0)
        item["yesterday_spend"] = float(item.get("yesterday_spend") or 0)
        for key in ("account_balance", "roi_day", "roi_week", "roi_iaa", "roi_ratio", "yesterday_roi_iaa"):
            value = item.get(key)
            item[key] = None if value is None else float(value)
        return item

    def _normalize_conditions(self, conditions):
        normalized = []
        field_keys = {item["key"] for item in FIELD_DEFS}
        for condition in conditions:
            field = str(condition.get("field") or "").strip()
            op = str(condition.get("op") or "").strip()
            if field not in field_keys or not op:
                continue
            normalized.append({"field": field, "op": op, "value": condition.get("value", "")})
        return normalized

    def _normalize_action(self, action):
        action_type = str(action.get("type") or "").strip()
        valid_actions = {item["key"] for item in ACTION_DEFS}
        if action_type not in valid_actions:
            raise ValueError("请选择有效动作")
        value = action.get("value")
        if action_type in VALUE_ACTIONS:
            mode = str(action.get("mode") or "set").strip()
            if mode not in ACTION_MODES:
                raise ValueError("请选择有效计算方式")
            if value in (None, ""):
                raise ValueError("动作参数不能为空")
            if to_number(value) is None:
                raise ValueError("动作参数必须是数字")
            normalized = {"type": action_type, "mode": mode, "value": value}
            max_value = action.get("max_value")
            if max_value not in (None, ""):
                max_number = to_number(max_value)
                if max_number is None:
                    raise ValueError("上限值必须是数字")
                if max_number <= 0:
                    raise ValueError("上限值必须大于 0")
                normalized["max_value"] = max_value
            return normalized
        return {"type": action_type}

    def _first_matching_rule(self, candidate, rules):
        for rule in rules:
            conditions = rule.get("conditions") or []
            if not conditions:
                continue
            results = [self._condition_matches(candidate, condition) for condition in conditions]
            if rule.get("match_type") == "any":
                if any(results):
                    return rule
            elif all(results):
                return rule
        return None

    def _condition_matches(self, candidate, condition):
        field = condition.get("field")
        op = condition.get("op")
        expected = condition.get("value")
        actual = candidate.get(field)
        if op == "is_empty":
            return actual in (None, "")
        if op == "not_empty":
            return actual not in (None, "")
        if op in {"contains", "not_contains"}:
            result = str(expected) in str(actual or "")
            return not result if op == "not_contains" else result
        if op in {"in", "not_in"}:
            values = [str(item).strip() for item in str(expected).replace("，", ",").split(",") if str(item).strip()]
            result = str(actual) in values
            return not result if op == "not_in" else result
        if op in {"eq", "ne"}:
            actual_number = to_number(actual)
            expected_number = to_number(expected)
            if actual_number is not None and expected_number is not None:
                result = actual_number == expected_number
            else:
                result = str(actual or "") == str(expected)
            return not result if op == "ne" else result

        actual_number = to_number(actual)
        expected_number = to_number(expected)
        if actual_number is None or expected_number is None:
            return False
        if op == "gt":
            return actual_number > expected_number
        if op == "gte":
            return actual_number >= expected_number
        if op == "lt":
            return actual_number < expected_number
        if op == "lte":
            return actual_number <= expected_number
        return False

    def _safety_check(self, rule, candidate, settings, run_counts, account_counts, allowed_account_ids=None):
        allowed_account_ids = allowed_account_ids if allowed_account_ids is not None else set(read_enabled_demand_account_ids())
        if candidate["advertiser_id"] not in allowed_account_ids:
            return "广告主不在需求账号.xlsx启用范围内"
        if candidate["advertiser_id"] in settings["blocked_advertiser_ids"]:
            return "广告主在黑名单中"
        if candidate["campaign_id"] in settings["blocked_campaign_ids"]:
            return "计划在黑名单中"
        if candidate["unit_id"] in settings["blocked_unit_ids"]:
            return "广告组在黑名单中"
        same_state = self._same_delivery_status_reason(rule.get("action") or {}, candidate)
        if same_state:
            return same_state
        max_run = settings["max_actions_per_run"]
        if max_run and run_counts["total"] >= max_run:
            return "达到单轮最大操作数"
        max_account = settings["max_actions_per_account_per_run"]
        account_count = account_counts.get(candidate["advertiser_id"], 0)
        if max_account and account_count >= max_account:
            return "达到单账号单轮最大操作数"
        max_rule_day = int(rule.get("max_daily_triggers") if rule.get("max_daily_triggers") is not None else 1)
        if max_rule_day and self._success_count_today(candidate["unit_id"], rule.get("rule_id")) >= max_rule_day:
            return f"达到本规则每日触发次数({max_rule_day})"
        return ""

    def _same_delivery_status_reason(self, action, candidate):
        action_type = action.get("type")
        if action_type not in {"pause_unit", "enable_unit"}:
            return ""
        current = to_number(candidate.get("put_status"))
        target = 2 if action_type == "pause_unit" else 1
        if current is not None and int(current) == target:
            return "当前投放状态已符合目标"
        return ""

    def _success_count_today(self, unit_id, rule_id=None):
        today = shanghai_today()
        params = [int(unit_id), today]
        rule_filter = ""
        if rule_id is not None:
            rule_filter = " AND rule_id = ?"
            params.append(int(rule_id))
        row = self.conn.execute(
            f"""
            SELECT COUNT(*) AS total
            FROM automation_logs
            WHERE unit_id = ?
              AND status = 'success'
              AND substr(created_at, 1, 10) = ?
              {rule_filter}
            """,
            params,
        ).fetchone()
        return int(row["total"] or 0)

    def _apply_action(self, action, candidate, dry_run):
        try:
            request = self._build_request(action, candidate)
        except ActionSkipped as skipped:
            return {"status": "skipped", "reason": str(skipped), "request": skipped.request, "response": {}}
        already = self._already_in_target_state(action, candidate, request["body"])
        if already:
            return {"status": "skipped", "reason": already, "request": request, "response": {}}
        if dry_run:
            return {"status": "dry_run", "reason": "试运行未调用快手接口", "request": request, "response": {}}

        try:
            response = self.client.request(request["endpoint"], method="POST", body=request["body"])
            self.client._assert_success(response)
            self._apply_local_success(action, candidate, request["body"])
            return {"status": "success", "reason": "快手接口返回成功", "request": request, "response": response}
        except KuaishouApiError as error:
            response = error.body if isinstance(error.body, dict) else {"error": str(error)}
            return {"status": "failed", "reason": str(error), "request": request, "response": response}
        except Exception as error:
            return {"status": "failed", "reason": str(error), "request": request, "response": {}}

    def _build_request(self, action, candidate):
        action_type = action["type"]
        advertiser_id = int(candidate["advertiser_id"])
        unit_id = int(candidate["unit_id"])
        if action_type in {"pause_unit", "enable_unit"}:
            put_status = 2 if action_type == "pause_unit" else 1
            return {
                "endpoint": UNIT_STATUS_ENDPOINT,
                "body": {"advertiser_id": advertiser_id, "unit_ids": [unit_id], "put_status": put_status},
            }
        if action_type == "set_roi_ratio":
            roi_ratio, calculation = self._calculate_action_target(action, candidate)
            roi_ratio = round(roi_ratio, 3)
            if roi_ratio <= 0 or roi_ratio > 100:
                raise ValueError("ROI系数必须在 0 到 100 之间")
            return {
                "endpoint": UNIT_UPDATE_ENDPOINT,
                "body": {"advertiser_id": advertiser_id, "unit_id": unit_id, "roi_ratio": roi_ratio},
                "calculation": calculation,
            }
        if action_type == "set_day_budget":
            day_budget_yuan, calculation = self._calculate_action_target(action, candidate)
            day_budget = money_yuan_to_li(day_budget_yuan)
            return {
                "endpoint": UNIT_BUDGET_ENDPOINT,
                "body": {"advertiser_id": advertiser_id, "unit_id": unit_id, "day_budget": day_budget},
                "calculation": calculation,
            }
        raise ValueError(f"不支持的动作: {action_type}")

    def _calculate_action_target(self, action, candidate):
        action_type = action["type"]
        mode = str(action.get("mode") or "set").strip()
        value = to_number(action.get("value"))
        if value is None:
            raise ValueError("动作参数必须是数字")
        max_value = to_number(action.get("max_value"))
        base = None
        base_field = None
        label = "当前ROI系数" if action_type == "set_roi_ratio" else "当前预算"
        if action_type == "set_roi_ratio":
            base_field = "roi_ratio"
            base = to_number(candidate.get("roi_ratio"))
        elif action_type == "set_day_budget":
            base_field = "day_budget_yuan"
            base = to_number(candidate.get("day_budget_yuan"))
            if base is None:
                base = li_to_yuan(candidate.get("day_budget"))

        if mode == "set":
            target = value
        elif mode in {"add_value", "add_percent"}:
            if base is None:
                raise ValueError(f"{label}为空，不能按当前值计算")
            target = base + value if mode == "add_value" else base * (1 + value / 100)
        else:
            raise ValueError("请选择有效计算方式")

        uncapped_target = target
        capped = False
        if max_value is not None:
            if base is not None and base >= max_value:
                calculation = {
                    "mode": mode,
                    "input_value": value,
                    "base_field": base_field,
                    "base_value": base,
                    "target_value": round(base, 3),
                    "uncapped_target_value": round(uncapped_target, 3),
                    "max_value": max_value,
                    "capped": False,
                }
                request = {"calculation": calculation}
                raise ActionSkipped(f"{label}已达到上限 {round(max_value, 3)}，不再修改", request=request)
            if target > max_value:
                target = max_value
                capped = True

        return target, {
            "mode": mode,
            "input_value": value,
            "base_field": base_field,
            "base_value": base,
            "target_value": round(target, 3),
            "uncapped_target_value": round(uncapped_target, 3),
            "max_value": max_value,
            "capped": capped,
        }

    def _already_in_target_state(self, action, candidate, body):
        action_type = action["type"]
        if action_type in {"pause_unit", "enable_unit"}:
            current = to_number(candidate.get("put_status"))
            target = body["put_status"]
            if current is not None and int(current) == target:
                return "当前投放状态已符合目标"
        if action_type == "set_roi_ratio":
            current = to_number(candidate.get("roi_ratio"))
            if current is not None and abs(current - float(body["roi_ratio"])) < 0.0005:
                return "当前ROI系数已符合目标"
        if action_type == "set_day_budget":
            current = to_number(candidate.get("day_budget"))
            if current is not None and int(round(current)) == int(body["day_budget"]):
                return "当前预算已符合目标"
        return ""

    def _apply_local_success(self, action, candidate, body):
        action_type = action["type"]
        updates = {"last_seen_at": local_now(), "last_sync_run_id": None}
        if action_type in {"pause_unit", "enable_unit"}:
            updates["put_status"] = int(body["put_status"])
        elif action_type == "set_roi_ratio":
            updates["roi_ratio"] = float(body["roi_ratio"])
        elif action_type == "set_day_budget":
            updates["day_budget"] = int(body["day_budget"])
        assignments = ", ".join(f"{key} = ?" for key in updates)
        values = list(updates.values())
        values.extend([int(candidate["advertiser_id"]), int(candidate["unit_id"])])
        self.conn.execute(
            f"UPDATE units SET {assignments} WHERE advertiser_id = ? AND unit_id = ?",
            values,
        )
        self.conn.commit()

    def _insert_log(self, batch_id, source, rule, candidate, status, reason, request, response):
        now = local_now()
        cursor = self.conn.execute(
            """
            INSERT INTO automation_logs(
              batch_id, source, rule_id, rule_name,
              advertiser_id, campaign_id, campaign_name,
              unit_id, unit_name, action_type, action_value_json,
              status, reason, request_json, response_json, before_json, created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                batch_id,
                source,
                rule.get("rule_id"),
                rule.get("name"),
                candidate.get("advertiser_id"),
                candidate.get("campaign_id"),
                candidate.get("campaign_name"),
                candidate.get("unit_id"),
                candidate.get("unit_name"),
                (rule.get("action") or {}).get("type"),
                json_text(rule.get("action") or {}),
                status,
                reason,
                json_text(request or {}),
                json_text(response or {}),
                json_text(candidate or {}),
                now,
            ),
        )
        self.conn.commit()
        return cursor.lastrowid
